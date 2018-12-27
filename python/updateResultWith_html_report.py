#!python3
# Usage: Update test result to assignment sheet after finish execution
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from os.path import relpath
import os, sys, fnmatch, codecs, time
from win32com.client import Dispatch

FUNCTION_NAME_COLUMN    ='B'
C0_COLUMN               ='BJ'
C1_COLUMN               ='BK'
MCDC_COLUMN             ='BL'
MATCH_COLUMN            =''
RESULT_COLUMN           ='BM'
EXECUTING_DATE_COLUMN   ='BO'
TASKSLIST               ='TasksList'
TEST_PREFIX             ='Test_'
# Pattern to get data in test result (.htm)
C0_Pattern      ='All C0 Coverage Rate'
C1_Pattern      ='All C1 Coverage Rate'
MCDC_Pattern    =''
Result_Pattern  ='All Output Comparison'
Match_Pattern   ='全体の一致'

def main():
    AssignmentFilePath='F:\\work\\3_5G_Soft_2\\branches\\eTAS\\03_Documents\\06_Assignment\\00_Detail_Assignment.xlsx'
    TestResultFolder = '\\\\GIT-011-064\\Test_Case_To_Execute\\Result'
    ErrorLogFolder='F:\work\\3_5G_Soft_2\\branches\\eTAS\\05_TestCases_Execution\\03_Errors_Notes'
#Parse htm
    #html_file = 'F:\\work\\3_5G_Soft_2\\branches\\2018_May_Release\\04_Output\\02_Test_Result\\KNL\\zynq\\cpu\\Test_GetCPUCoreID\\テスト結果報告書.htm'
    #html_file = 'F:\\work\\3_5G_Soft_2\\branches\\eTAS\\04_Output\\02_Test_Result\\Application\\APL_Etas\\Test_Rte_Write_Etas_StabilityJudgeStatus\\テスト結果報告書.htm'
    start_time = time.time()
    ResultDict = GetResult(TestResultFolder,ErrorLogFolder)
    UpdateResult2Excel(AssignmentFilePath,ResultDict)
    print("--- %s seconds ---" % (time.time() - start_time))

def GetResult(TestResultFolder,ErrorLogFolder):
    ErrorLogFile ='\\\\GIT-011-064\\Test_Case_To_Execute\Result\\Error_Log.txt'
    if not ErrorLogFolder: TestResultFolder
    ResultDict={}
    for root, dirs, files in os.walk(TestResultFolder):
        C0 = 'N/A'
        C1 = 'N/A'
        Mcdc = 'N/A'
        Result = 'N/A'
        Match = 'N/A'
        ExecutionTime = 0
        ResultPath=''
        isFolderTestReport = False
        for basename in files:
            # Get folder dir test result, test case ID, time execution. Because the file "TestResultInfo.data is always generated even there is error"
            # So it should be used to get test cases name, time execution.
            if fnmatch.fnmatch(basename, 'TestResultInfo.dat'):
                isFolderTestReport = True
                ResultPath=root
                tcID = os.path.basename(root)
                ExecutionTime= os.path.getmtime(os.path.join(root, basename))
                if Result == 'N/A': Result = 'Error'
                #executionTime= time.strftime("%d/%m %H:%M",time.gmtime(os.path.getctime(os.path.join(root, basename))))
            # Get test result from htm file
            if fnmatch.fnmatch(basename, '*.htm'):
                # Parse htm file to get result
                html_file = os.path.join(root,basename)
                (C0, C1, Mcdc,Result,Match)=ParsingTestReport(html_file)
        if isFolderTestReport == True:
            tcID = tcID.replace(TEST_PREFIX,'')
            #Find error log for test error
            if Result == 'Error':
                ErrorLog = find_files(ErrorLogFolder, '[A-Za-z0-9_]*' + tcID+'.txt')
                if len(ErrorLog) == 0:
                     print('WARNING: There is no error log for the test case ' + tcID)
                elif len(ErrorLog) == 1:
                     ResultPath = ErrorLog[0]
                else:
                     print ('ERROR: There are more than one error log for the test case ' + tcID)
                     ResultPath = ErrorLog[0]
                ResultPath =ErrorLogFile
            if tcID in ResultDict:
                print ('ERROR: There is more than one test result for the test case ' + tcID)
                if(ExecutionTime > ResultDict[tcID].execution_time):
                    # Replace latest result
                    ResultDict[tcID] = TestResult(C0, C1, Mcdc, Result, Match, ExecutionTime, ResultPath)
            else:
                ResultDict[tcID] = TestResult(C0, C1, Mcdc, Result, Match, ExecutionTime, ResultPath)
    return ResultDict
# Load excel file
def UpdateResult2Excel(AssignmentFilePath,ResultDict):
    AssingmentFolderPath = os.path.dirname(AssignmentFilePath)
    wb=load_workbook(AssignmentFilePath)
    ws=wb[TASKSLIST]
    # Get List of functions
    functionList=[]
    for cell in ws[FUNCTION_NAME_COLUMN]:
        functionList.append(cell.value)
    for key in ResultDict:
        try:
            RowIndex = functionList.index(key) + 1
            ws[C0_COLUMN + str(RowIndex)].value = ResultDict[key].c0
            ws[C1_COLUMN + str(RowIndex)].value = ResultDict[key].c0
            ws[RESULT_COLUMN + str(RowIndex)].value = ResultDict[key].result
            ws[EXECUTING_DATE_COLUMN + str(RowIndex)].value = time.ctime(ResultDict[key].execution_time)
            #ws[RESULT_COLUMN + str(RowIndex)].hyperlink = ResultDict[key].resultPath
            try:
                ws[RESULT_COLUMN + str(RowIndex)].hyperlink = relpath(ResultDict[key].resultPath,AssingmentFolderPath)
            except ValueError:
                print('WARNING: Can not create relative path for linking to the test result of the test case ' + key)
                ws[RESULT_COLUMN + str(RowIndex)].hyperlink = ResultDict[key].resultPath
                pass
            ws[MCDC_COLUMN + str(RowIndex)].value = ResultDict[key].mcdc
            ws[MATCH_COLUMN + str(RowIndex)].value = ResultDict[key].match
        except ValueError:
            print ("ERROR: The function " + key + " does not exit in the assignment file")
        except AttributeError:
            pass
    try:
        wb.save(AssignmentFilePath)
    except PermissionError:
        xl = Dispatch('Excel.Application')
        xl_wb=xl.Workbooks(os.path.basename(AssignmentFilePath))
        xl_wb.Close(False)
        wb.save(AssignmentFilePath)
        xl_wb = xl.Workbooks.Open(AssignmentFilePath)
        del(xl)
        del(xl_wb)
    print (len(ResultDict))
    print (len(functionList))

def ParsingTestReport(htmFullPath):
    # parsing htm to get data of first table
    with codecs.open(htmFullPath,'r',encoding='Shift_JIS') as f:
        text = f.read()
    soup=BeautifulSoup(text,'html.parser')
    table=soup.find("table")
    listStr= table.getText(separator='\n',strip=True).split('\n')
    try:
        C0_Value = listStr[1 + listStr.index(C0_Pattern)]
    except ValueError:
        C0_Value = 'N/A'
    try:
        C1_Value = listStr[1 + listStr.index(C1_Pattern)]
    except ValueError:
        C1_Value = 'N/A'
    try:
        MCDC_Value = listStr[1 + listStr.index(MCDC_Pattern)]
    except ValueError:
        MCDC_Value = 'N/A'
    try:
        Result_Value = listStr[1 + listStr.index(Result_Pattern)]
    except ValueError:
        Result_Value = 'N/A'
    try:
        Match_Value =  listStr[1 + listStr.index(Match_Pattern)]
    except ValueError:
        Match_Value = 'N/A'

    return (C0_Value, C1_Value, MCDC_Value, Result_Value,Match_Value)

def find_files(directory, pattern):
    fileList=[]
    for root, dirs, files in os.walk(directory):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                fileList.append(filename)
    return fileList

class TestResult:
    def __init__(self, c0, c1, mcdc, result,match, execution_time, resultPath):
        self.c0 = c0
        self.c1 = c1
        self.mcdc = mcdc
        self.result = result
        self.match = match
        self.execution_time = execution_time
        self.resultPath = resultPath

if __name__ == '__main__':
    main()
